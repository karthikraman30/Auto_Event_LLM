# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html

from itemadapter import ItemAdapter
from openpyxl import Workbook


class EventCategoryPipeline:
    def process_item(self, item, spider):
        return item


class ExcelExportPipeline:
    """Pipeline to export items to Excel file, sorted by date."""
    
    def __init__(self):
        self.items = []  # Collect all items first
    
    def open_spider(self, spider):
        pass  # Nothing to do on open
    
    def close_spider(self, spider):
        # Sort items by date_iso
        self.items.sort(key=lambda x: ItemAdapter(x).get("date_iso", "") or "9999-99-99")
        
        # Create workbook and write sorted items
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Events"
        
        # Add headers
        headers = [
            "Event Name", "Date", "Date ISO", "End Date ISO", "Time", "Location", 
            "Target Group", "Target Group Normalized", "Status",
            "Description", "Event URL"
        ]
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Write sorted items
        for row, item in enumerate(self.items, 2):
            adapter = ItemAdapter(item)
            worksheet.cell(row=row, column=1, value=adapter.get("event_name", ""))
            worksheet.cell(row=row, column=2, value=adapter.get("date", ""))
            worksheet.cell(row=row, column=3, value=adapter.get("date_iso", ""))
            worksheet.cell(row=row, column=4, value=adapter.get("end_date_iso", "N/A"))
            worksheet.cell(row=row, column=5, value=adapter.get("time", ""))
            worksheet.cell(row=row, column=6, value=adapter.get("location", ""))
            worksheet.cell(row=row, column=7, value=adapter.get("target_group", ""))
            worksheet.cell(row=row, column=8, value=adapter.get("target_group_normalized", ""))
            worksheet.cell(row=row, column=9, value=adapter.get("status", ""))
            worksheet.cell(row=row, column=10, value=adapter.get("description", ""))
            worksheet.cell(row=row, column=11, value=adapter.get("event_url", ""))
        
        # Save the workbook
        filename = "events.xlsx"
        workbook.save(filename)
        spider.logger.info(f"Exported {len(self.items)} events to {filename} (sorted by date)")
    
    def process_item(self, item, spider):
        self.items.append(item)
        return item
