# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html

from itemadapter import ItemAdapter
from openpyxl import Workbook


class EventCategoryPipeline:
    def process_item(self, item, spider):
        return item


class ExcelExportPipeline:
    """Pipeline to export items to Excel file, sorted by date, with dynamic columns."""
    
    def __init__(self):
        self.items = []  # Collect all items first
    
    def open_spider(self, spider):
        pass  # Nothing to do on open
    
    def close_spider(self, spider):
        # 1. Sort items by date_iso
        self.items.sort(key=lambda x: ItemAdapter(x).get("date_iso", "") or "9999-99-99")
        
        # 2. Collect all unique keys from 'extra_attributes' across ALL items
        # This ensures that if one event has "Price" and another has "Speaker", we get columns for both.
        dynamic_keys = set()
        for item in self.items:
            adapter = ItemAdapter(item)
            extras = adapter.get("extra_attributes", {})
            if extras:
                dynamic_keys.update(extras.keys())
        
        # Sort them so columns are always in the same order
        sorted_dynamic_keys = sorted(list(dynamic_keys))

        # 3. Create workbook
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Events"
        
        # 4. Define Standard Headers
        fixed_headers = [
            "Event Name", "Date", "Date ISO", "End Date ISO", "Time", "Location", 
            "Target Group", "Target Group Normalized", "Status", 
            "Booking Info", "Description", "Event URL"
        ]
        
        # Combine Fixed Headers + Dynamic Headers
        all_headers = fixed_headers + sorted_dynamic_keys
        
        # Write Headers
        for col, header in enumerate(all_headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # 5. Write Data Rows
        for row_idx, item in enumerate(self.items, 2):
            adapter = ItemAdapter(item)
            
            # Write Fixed Columns (1 to 11)
            worksheet.cell(row=row_idx, column=1, value=adapter.get("event_name", ""))
            worksheet.cell(row=row_idx, column=2, value=adapter.get("date", ""))
            worksheet.cell(row=row_idx, column=3, value=adapter.get("date_iso", ""))
            worksheet.cell(row=row_idx, column=4, value=adapter.get("end_date_iso", "N/A"))
            worksheet.cell(row=row_idx, column=5, value=adapter.get("time", ""))
            worksheet.cell(row=row_idx, column=6, value=adapter.get("location", ""))
            worksheet.cell(row=row_idx, column=7, value=adapter.get("target_group", ""))
            worksheet.cell(row=row_idx, column=8, value=adapter.get("target_group_normalized", ""))
            worksheet.cell(row=row_idx, column=9, value=adapter.get("status", ""))
            worksheet.cell(row=row_idx, column=10, value=adapter.get("booking_info", "N/A"))
            worksheet.cell(row=row_idx, column=11, value=adapter.get("description", ""))
            worksheet.cell(row=row_idx, column=12, value=adapter.get("event_url", ""))
            
            
            # Write Dynamic Columns (12 onwards)
            # We look up the value in 'extra_attributes'. If not found, write empty string.
            extras = adapter.get("extra_attributes", {})
            for i, key in enumerate(sorted_dynamic_keys):
                col_num = len(fixed_headers) + 1 + i
                value = extras.get(key, "")
                worksheet.cell(row=row_idx, column=col_num, value=str(value)) # Ensure it's string
        
        # Save the workbook
        filename = "events.xlsx"
        workbook.save(filename)
        spider.logger.info(f"Exported {len(self.items)} events to {filename} with {len(sorted_dynamic_keys)} dynamic columns.")
    
    def process_item(self, item, spider):
        self.items.append(item)
        return item